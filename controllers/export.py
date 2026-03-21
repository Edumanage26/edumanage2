from flask import (Blueprint, request, session,
                   redirect, url_for, flash, send_file)
from database import get_db, close_db
from utils.decorators import login_required, roles_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import date

export_bp = Blueprint("export", __name__)

PRIMARY = "4F46E5"
SUCCESS = "059669"
WARNING = "D97706"
DANGER  = "DC2626"
DARK    = "0F172A"
WHITE   = "FFFFFF"


def thin_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(cell, bg=PRIMARY, fc=WHITE, size=11, bold=True):
    cell.font      = Font(bold=bold, color=fc, size=size, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()


def dat(cell, bg=WHITE, bold=False, align="left"):
    cell.font      = Font(size=10, name="Calibri", bold=bold)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = thin_border()


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def send_wb(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def add_title(ws, title, subtitle, cols):
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=cols)
    t = ws.cell(1, 1, title)
    t.font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    t.fill = PatternFill("solid", fgColor=DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(
        start_row=2, start_column=1, end_row=2, end_column=cols)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(size=11, color="64748B", name="Calibri")
    s.fill = PatternFill("solid", fgColor="F1F5F9")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20


@export_bp.route("/export/students")
@login_required
def export_students():
    school_id = session.get("school_id")
    conn = get_db()
    cur  = conn.cursor()

    if session.get("role") == "super_admin":
        cur.execute("""
            SELECT s.*, c.name AS class_name,
                   sc.name AS school_name
            FROM students s
            LEFT JOIN classes c ON c.id = s.class_id
            LEFT JOIN schools sc ON sc.id = s.school_id
            WHERE s.is_active = 1
            ORDER BY sc.name, c.name, s.first_name
        """)
    else:
        cur.execute("""
            SELECT s.*, c.name AS class_name
            FROM students s
            LEFT JOIN classes c ON c.id = s.class_id
            WHERE s.school_id = ? AND s.is_active = 1
            ORDER BY c.name, s.first_name
        """, (school_id,))
    students = cur.fetchall()

    cur.execute(
        "SELECT name FROM schools WHERE id = ?", (school_id,))
    sch = cur.fetchone()
    school_name = sch["name"] if sch else "School"
    close_db(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    add_title(ws, f"{school_name} - Student List",
              f"Exported on {date.today().strftime('%d %B %Y')}", 8)

    for i, h in enumerate(["#", "Full Name", "Admission No.",
                            "Class", "Gender", "Date of Birth",
                            "Parent Phone", "Status"], 1):
        hdr(ws.cell(3, i, h))
    ws.row_dimensions[3].height = 20

    for r, s in enumerate(students, 1):
        row = r + 3
        bg  = WHITE if r % 2 == 0 else "F8FAFC"
        vals = [r, f"{s['first_name']} {s['last_name']}",
                s["admission_no"] or "---",
                s["class_name"] or "Unassigned",
                s["gender"] or "---",
                s["date_of_birth"] or "---",
                s["parent_phone"] or "---", "Active"]
        for i, v in enumerate(vals, 1):
            dat(ws.cell(row, i, v), bg=bg, bold=(i == 2),
                align="center" if i in [1, 5, 8] else "left")
        ws.row_dimensions[row].height = 16

    set_col_widths(ws, [5, 25, 16, 14, 10, 14, 16, 10])
    ws.freeze_panes = "A4"
    return send_wb(wb, f"students_{date.today()}.xlsx")


@export_bp.route("/export/attendance")
@login_required
def export_attendance():
    school_id  = session.get("school_id")
    class_id   = request.args.get("class_id", "")
    start_date = request.args.get("start_date", "")
    end_date   = request.args.get("end_date", "")
    conn = get_db()
    cur  = conn.cursor()

    query = """
        SELECT a.date, a.status,
               s.first_name, s.last_name, s.admission_no,
               c.name AS class_name
        FROM attendance a
        JOIN students s ON s.id = a.student_id
        JOIN classes c ON c.id = a.class_id
        WHERE a.school_id = ?
    """
    params = [school_id]
    if class_id:
        query += " AND a.class_id = ?"
        params.append(class_id)
    if start_date:
        query += " AND a.date >= ?"
        params.append(start_date)
    if end_date:
        query += " AND a.date <= ?"
        params.append(end_date)
    query += " ORDER BY a.date DESC, c.name, s.first_name"
    cur.execute(query, params)
    records = cur.fetchall()

    cur.execute(
        "SELECT name FROM schools WHERE id = ?", (school_id,))
    sch = cur.fetchone()
    school_name = sch["name"] if sch else "School"
    close_db(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    add_title(ws, f"{school_name} - Attendance Records",
              f"Exported on {date.today().strftime('%d %B %Y')}", 6)

    for i, h in enumerate(["#", "Date", "Student Name",
                            "Admission No.", "Class", "Status"], 1):
        hdr(ws.cell(3, i, h))
    ws.row_dimensions[3].height = 20

    for r, rec in enumerate(records, 1):
        row = r + 3
        bg  = WHITE if r % 2 == 0 else "F8FAFC"
        vals = [r, rec["date"],
                f"{rec['first_name']} {rec['last_name']}",
                rec["admission_no"] or "---",
                rec["class_name"], rec["status"]]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v)
            dat(c, bg=bg,
                align="center" if i in [1, 2, 5] else "left")
            if i == 6:
                c.font = Font(
                    bold=True, size=10, name="Calibri",
                    color=SUCCESS if v == "Present" else DANGER)
        ws.row_dimensions[row].height = 16

    set_col_widths(ws, [5, 14, 25, 16, 14, 12])
    ws.freeze_panes = "A4"
    return send_wb(wb, f"attendance_{date.today()}.xlsx")


@export_bp.route("/export/results")
@login_required
def export_results():
    school_id  = session.get("school_id")
    class_id   = request.args.get("class_id", "")
    term       = request.args.get("term", "First Term")
    session_yr = request.args.get("session", "2024/2025")

    if not class_id:
        flash("Please select a class.", "warning")
        return redirect(url_for("result.index"))

    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        SELECT s.first_name, s.last_name, s.admission_no,
               sub.name AS subject,
               r.ca1, r.exam, r.score, r.grade
        FROM results r
        JOIN students s ON s.id = r.student_id
        JOIN subjects sub ON sub.id = r.subject_id
        WHERE r.class_id = ? AND r.term = ? AND r.session = ?
        ORDER BY s.first_name, sub.name
    """, (class_id, term, session_yr))
    records = cur.fetchall()

    cur.execute(
        "SELECT name FROM classes WHERE id = ?", (class_id,))
    cls = cur.fetchone()
    class_name = cls["name"] if cls else "Class"

    cur.execute(
        "SELECT name FROM schools WHERE id = ?", (school_id,))
    sch = cur.fetchone()
    school_name = sch["name"] if sch else "School"
    close_db(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    add_title(
        ws, f"{school_name} - {class_name} Results",
        f"{term} | {session_yr} | Exported {date.today().strftime('%d %B %Y')}",
        8)

    for i, h in enumerate(["#", "Student Name", "Admission No.",
                            "Subject", "CA (/40)", "Exam (/60)",
                            "Total (/100)", "Grade"], 1):
        hdr(ws.cell(3, i, h))
    ws.row_dimensions[3].height = 20

    for r, rec in enumerate(records, 1):
        row = r + 3
        bg  = WHITE if r % 2 == 0 else "F8FAFC"
        vals = [r, f"{rec['first_name']} {rec['last_name']}",
                rec["admission_no"] or "---", rec["subject"],
                rec["ca1"] or 0, rec["exam"] or 0,
                rec["score"] or 0, rec["grade"] or "---"]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v)
            dat(c, bg=bg, bold=(i in [2, 7]),
                align="center" if i in [1, 5, 6, 7, 8] else "left")
            if i == 8:
                if v == "A":
                    c.font = Font(bold=True, color=SUCCESS,
                                  size=10, name="Calibri")
                elif v == "F":
                    c.font = Font(bold=True, color=DANGER,
                                  size=10, name="Calibri")
        ws.row_dimensions[row].height = 16

    set_col_widths(ws, [5, 25, 16, 20, 10, 10, 12, 8])
    ws.freeze_panes = "A4"
    return send_wb(
        wb, f"results_{class_name}_{term}_{date.today()}.xlsx")


@export_bp.route("/export/fees")
@login_required
@roles_required("super_admin", "school_admin", "accounts_officer")
def export_fees():
    school_id  = session.get("school_id")
    class_id   = request.args.get("class_id", "")
    term       = request.args.get("term", "First Term")
    session_yr = request.args.get("session", "2024/2025")

    if not class_id:
        flash("Please select a class.", "warning")
        return redirect(url_for("fee.report"))

    conn = get_db()
    cur  = conn.cursor()

    cur.execute(
        "SELECT amount FROM fee_structure WHERE class_id = ? AND term = ? AND session = ?",
        (class_id, term, session_yr))
    fs = cur.fetchone()
    fee_amount = fs["amount"] if fs else 0

    cur.execute("""
        SELECT s.first_name, s.last_name, s.admission_no,
               COALESCE(SUM(fp.amount_paid), 0) AS amount_paid
        FROM students s
        LEFT JOIN fee_payments fp ON fp.student_id = s.id
        WHERE s.class_id = ? AND s.is_active = 1
        GROUP BY s.id ORDER BY s.first_name
    """, (class_id,))
    students = cur.fetchall()

    cur.execute(
        "SELECT name FROM classes WHERE id = ?", (class_id,))
    cls = cur.fetchone()
    class_name = cls["name"] if cls else "Class"

    cur.execute(
        "SELECT name FROM schools WHERE id = ?", (school_id,))
    sch = cur.fetchone()
    school_name = sch["name"] if sch else "School"
    close_db(conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fee Report"
    add_title(
        ws, f"{school_name} - {class_name} Fee Report",
        f"{term} | {session_yr} | Exported {date.today().strftime('%d %B %Y')}",
        7)

    for i, h in enumerate(["#", "Student Name", "Admission No.",
                            "Fee Amount", "Amount Paid",
                            "Balance", "Status"], 1):
        hdr(ws.cell(3, i, h), bg=SUCCESS)
    ws.row_dimensions[3].height = 20

    for r, s in enumerate(students, 1):
        row     = r + 3
        bg      = WHITE if r % 2 == 0 else "F8FAFC"
        balance = max(0, fee_amount - s["amount_paid"])
        status  = ("Paid" if balance <= 0
                   else "Partial" if s["amount_paid"] > 0
                   else "Unpaid")
        vals = [r, f"{s['first_name']} {s['last_name']}",
                s["admission_no"] or "---",
                f"N{fee_amount:,.2f}",
                f"N{s['amount_paid']:,.2f}",
                f"N{balance:,.2f}", status]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row, i, v)
            dat(c, bg=bg, bold=(i == 2),
                align="center" if i in [1, 3, 7]
                else "right" if i in [4, 5, 6] else "left")
            if i == 7:
                if v == "Paid":
                    c.font = Font(bold=True, color=SUCCESS,
                                  size=10, name="Calibri")
                elif v == "Partial":
                    c.font = Font(bold=True, color=WARNING,
                                  size=10, name="Calibri")
                else:
                    c.font = Font(bold=True, color=DANGER,
                                  size=10, name="Calibri")
        ws.row_dimensions[row].height = 16

    set_col_widths(ws, [5, 25, 16, 14, 14, 14, 10])
    ws.freeze_panes = "A4"
    return send_wb(
        wb, f"fees_{class_name}_{term}_{date.today()}.xlsx")


@export_bp.route("/export/reportcard")
@login_required
def export_reportcard():
    school_id  = session.get("school_id")
    class_id   = request.args.get("class_id", "")
    term       = request.args.get("term", "First Term")
    session_yr = request.args.get("session", "2024/2025")

    if not class_id:
        flash("Please select a class.", "warning")
        return redirect(url_for("result.index"))

    conn = get_db()
    cur  = conn.cursor()

    cur.execute("""
        SELECT c.*, sc.name AS school_name
        FROM classes c
        JOIN schools sc ON sc.id = c.school_id
        WHERE c.id = ?
    """, (class_id,))
    cls = cur.fetchone()

    cur.execute(
        "SELECT * FROM subjects WHERE class_id = ? ORDER BY name",
        (class_id,))
    subjects = cur.fetchall()

    cur.execute("""
        SELECT * FROM students
        WHERE class_id = ? AND is_active = 1
        ORDER BY first_name
    """, (class_id,))
    students = cur.fetchall()

    cur.execute(
        "SELECT COUNT(*) AS cnt FROM students WHERE class_id = ? AND is_active = 1",
        (class_id,))
    total_students = cur.fetchone()["cnt"]

    class_name  = cls["name"]        if cls else "Class"
    school_name = cls["school_name"] if cls else "School"

    all_totals  = []
    report_data = []

    for s in students:
        subj_rows   = []
        total_score = 0
        count       = 0

        for sub in subjects:
            cur.execute("""
                SELECT ca1, exam, score, grade FROM results
                WHERE student_id = ? AND subject_id = ?
                AND term = ? AND session = ?
            """, (s["id"], sub["id"], term, session_yr))
            row = cur.fetchone()
            if row and row["score"] is not None:
                total_score += row["score"]
                count += 1
                subj_rows.append({
                    "subject": sub["name"],
                    "ca1":    row["ca1"]   or 0,
                    "exam":   row["exam"]  or 0,
                    "total":  row["score"] or 0,
                    "grade":  row["grade"] or "---",
                    "status": "Pass" if row["score"] >= 40 else "Fail"
                })
            else:
                subj_rows.append({
                    "subject": sub["name"],
                    "ca1": 0, "exam": 0,
                    "total": 0, "grade": "---", "status": "---"
                })

        average    = round(total_score / count, 1) if count else 0
        percentage = round(
            (total_score / (count * 100)) * 100, 1) if count else 0
        all_totals.append({"id": s["id"], "total": total_score})

        cur.execute("""
            SELECT COUNT(*) AS total,
                SUM(CASE WHEN status = 'Present' THEN 1 ELSE 0 END) AS present,
                SUM(CASE WHEN status = 'Absent'  THEN 1 ELSE 0 END) AS absent
            FROM attendance
            WHERE student_id = ? AND class_id = ?
        """, (s["id"], class_id))
        att = cur.fetchone()

        if average >= 80:
            comment = "Excellent performance! Keep up the outstanding work."
        elif average >= 70:
            comment = "Very good performance. Continue to put in great effort."
        elif average >= 60:
            comment = "Good performance. There is room for improvement."
        elif average >= 50:
            comment = "Average performance. More effort is needed."
        elif average >= 40:
            comment = "Below average. Needs to work harder."
        else:
            comment = "Poor performance. Urgent attention required."

        report_data.append({
            "student":     s,
            "subjects":    subj_rows,
            "total_score": round(total_score, 1),
            "average":     average,
            "percentage":  percentage,
            "comment":     comment,
            "att_total":   att["total"]   or 0 if att else 0,
            "att_present": att["present"] or 0 if att else 0,
            "att_absent":  att["absent"]  or 0 if att else 0,
        })

    sorted_totals = sorted(
        all_totals, key=lambda x: x["total"], reverse=True)
    positions = {item["id"]: i+1
                 for i, item in enumerate(sorted_totals)}
    close_db(conn)

    wb = Workbook()
    wb.remove(wb.active)

    for data in report_data:
        s      = data["student"]
        pos    = positions[s["id"]]
        suffix = ("st" if pos == 1 else "nd" if pos == 2
                  else "rd" if pos == 3 else "th")
        att_rate = (
            f"{int((data['att_present']/data['att_total'])*100)}%"
            if data["att_total"] > 0 else "N/A")
        name = f"{s['first_name']} {s['last_name']}"
        ws   = wb.create_sheet(title=name[:31])

        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value     = f"{school_name.upper()} - STUDENT REPORT CARD"
        c.font      = Font(bold=True, size=14, color=WHITE,
                           name="Calibri")
        c.fill      = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(
            horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:H2")
        c = ws["A2"]
        c.value = (f"{term.upper()}  |  {session_yr}  |  "
                   f"Exported: {date.today().strftime('%d %B %Y')}")
        c.font      = Font(bold=True, size=11, color=WHITE,
                           name="Calibri")
        c.fill      = PatternFill("solid", fgColor=PRIMARY)
        c.alignment = Alignment(
            horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 8

        info_rows = [
            ("Full Name",        name),
            ("Admission Number", s["admission_no"] or "---"),
            ("Class",            class_name),
            ("Gender",           s["gender"] or "---"),
            ("Number in Class",  str(total_students)),
            ("Class Position",   f"{pos}{suffix} of {total_students}"),
            ("Total Score",      str(data["total_score"])),
            ("Average Score",    str(data["average"])),
            ("Percentage",       f"{data['percentage']}%"),
        ]
        row = 4
        for label, value in info_rows:
            lc = ws.cell(row, 1, label)
            lc.font      = Font(bold=True, size=10, name="Calibri")
            lc.fill      = PatternFill("solid", fgColor="F1F5F9")
            lc.alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            lc.border    = thin_border()
            vc = ws.cell(row, 2, value)
            vc.font      = Font(size=10, name="Calibri",
                                bold=True, color=PRIMARY)
            vc.fill      = PatternFill("solid", fgColor="EEF2FF")
            vc.alignment = Alignment(
                horizontal="left", vertical="center", indent=1)
            vc.border    = thin_border()
            ws.merge_cells(
                start_row=row, start_column=3,
                end_row=row, end_column=8)
            blank = ws.cell(row, 3, "")
            blank.fill   = PatternFill("solid", fgColor=WHITE)
            blank.border = thin_border()
            ws.row_dimensions[row].height = 18
            row += 1

        row += 2
        for i, h in enumerate(["#", "Subject", "CA (/40)",
                                "Exam (/60)", "Total (/100)",
                                "Grade", "Status", ""], 1):
            hdr(ws.cell(row, i, h))
        ws.row_dimensions[row].height = 20
        row += 1

        for idx, sub in enumerate(data["subjects"], 1):
            bg   = WHITE if idx % 2 == 0 else "F8FAFC"
            vals = [idx, sub["subject"], sub["ca1"],
                    sub["exam"], sub["total"],
                    sub["grade"], sub["status"], ""]
            for i, v in enumerate(vals, 1):
                c = ws.cell(row, i, v)
                dat(c, bg=bg, bold=(i in [2, 5]),
                    align="center" if i in [1,3,4,5,6,7]
                    else "left")
                if i == 6:
                    if v == "A":
                        c.font = Font(bold=True, color=SUCCESS,
                                      size=10, name="Calibri")
                    elif v == "F":
                        c.font = Font(bold=True, color=DANGER,
                                      size=10, name="Calibri")
                if i == 7:
                    if v == "Pass":
                        c.font = Font(bold=True, color=SUCCESS,
                                      size=10, name="Calibri")
                    elif v == "Fail":
                        c.font = Font(bold=True, color=DANGER,
                                      size=10, name="Calibri")
            ws.row_dimensions[row].height = 16
            row += 1

        row += 2
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=8)
        ac = ws.cell(row, 1,
            f"ATTENDANCE  |  Total Days: {data['att_total']}  |  "
            f"Present: {data['att_present']}  |  "
            f"Absent: {data['att_absent']}  |  Rate: {att_rate}")
        ac.font      = Font(bold=True, size=10, name="Calibri",
                            color="0C4A6E")
        ac.fill      = PatternFill("solid", fgColor="F0F9FF")
        ac.alignment = Alignment(
            horizontal="left", vertical="center", indent=1)
        ac.border    = thin_border()
        ws.row_dimensions[row].height = 20
        row += 2

        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row, end_column=8)
        cc = ws.cell(row, 1,
            f"TEACHER'S COMMENT:  {data['comment']}")
        cc.font      = Font(size=10, italic=True, name="Calibri",
                            color="92400E")
        cc.fill      = PatternFill("solid", fgColor="FFFBEB")
        cc.alignment = Alignment(
            horizontal="left", vertical="center",
            wrap_text=True, indent=1)
        cc.border    = thin_border()
        ws.row_dimensions[row].height = 30

        set_col_widths(ws, [20, 25, 12, 12, 14, 10, 10, 5])
        ws.freeze_panes = "A4"

    fname = f"reportcard_{class_name}_{term}_{date.today()}.xlsx"
    return send_wb(wb, fname)
